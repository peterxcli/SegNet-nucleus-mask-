import torch
from torch import nn
from torch import random
from torch.utils.data import Dataset, DataLoader
from os import listdir
import numpy as np
import cv2 as cv
import torch.nn.functional as F

from main import NucleusDataset, SegNet, Fmeasure, MAE

import numpy as np
import matplotlib.pyplot as plt
from matplotlib.colors import ListedColormap
from torch.autograd import Variable
import random
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, colors, Alignment


def preprocess(epoch, index):
    test_dataset = NucleusDataset("test")
    foldername = test_dataset.folderlist[index]
    filename = foldername.split(".")[0] + ".png"
    img = cv.imread(test_dataset.root + foldername + "/images/" + filename)
    img = cv.resize(img, (224, 224), interpolation=cv.INTER_LINEAR)
    _input, obj = test_dataset.__getitem__(index)
    result = np.empty(shape=(224, 224))
    for x in range(224):
        for y in range(224):
            if obj[0][x][y] == 1:
                result[x][y] = 255
            else:
                result[x][y] = 0
    model = torch.load("Model/SegNet_bce_e{}.sav".format(epoch))
    model.cpu()
    model.eval()

    tmp = []
    tmp.append(_input)
    _input = tmp
    _input = Variable(torch.Tensor(_input).float())
    _output = model(_input)[0]
    _output = 1*(_output[:, :, :] > 0.5)
    output = np.empty(shape=(224, 224))
    for x in range(224):
        for y in range(224):
            if _output[0][x][y] == 1:
                output[x][y] = 255
            else:
                output[x][y] = 0
    return img, result, output  # ori, label, output


def test(n_samples):
    batch_size = 10
    test_dataset = NucleusDataset("test")
    index = random.randint(0, int(len(test_dataset.allfolderlist) / 2)-1)
    i = 0
    diff = 10
    demo = []
    while i < n_samples*diff:
        i += diff
        demo.append(preprocess(i, index))
    fig , ax = plt.subplots(3, n_samples, sharex=True, sharey=True)
    for i in range(3):
        for j in range(n_samples):
            ax[i, j].set_yticklabels([])
            ax[i, j].set_xticklabels([])
            if i != 0:
                ax[i, j].imshow(demo[j][i], cmap='gray', vmin = 0, vmax = 255, interpolation='none')
            else:
                ax[i, j].imshow(demo[j][i])

    plt.setp(ax[0, 0], ylabel='ori')
    plt.setp(ax[1, 0], ylabel='label')
    plt.setp(ax[2, 0], ylabel='output')

    for i in(0, n_samples-1):
        plt.setp(ax[2, i], xlabel='Epoch{}'.format((i+1)*diff))

    plt.savefig("test result.png")
    plt.close()

    wb = Workbook()
    sheet = wb.active
    sheet['A1'] = "Epoch"
    sheet['B1'] = "Normal"
    sheet['B2'] = "F-measure"
    sheet['C2'] = "MAE"
    sheet.merge_cells('B1:C1')
    sheet.merge_cells('A1:A2')
    test_loader = DataLoader(test_dataset, batch_size=batch_size, num_workers=1, drop_last=True, shuffle=False)
    for i in range(0, n_samples):
        sheet['A{}'.format(i+3)] = (i+1)*diff
        model = torch.load("Model/SegNet_bce_e{}.sav".format((i+1)*diff))
        model.eval()
        model.cpu()
        mae = 0
        fmeasure = 0
        for x, y in test_loader:
            x = x.float()
            x.cpu()
            y = y.float()
            y.cpu()
            output = model(x)
            mae = MAE(output, y, batch_size)
            fmeasure = Fmeasure(output, y, bs=batch_size)
            break
        sheet['B{}'.format(i+3)] = fmeasure
        sheet['C{}'.format(i+3)] = mae
    wb.save("test result.xlsx")


if __name__ == '__main__':
    test(2)
